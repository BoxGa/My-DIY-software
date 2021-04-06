# -*- coding: utf-8 -*-
# 以后代码前面要加上utf-8才不会出问题
import keyboard
import time
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import datetime

# timing1.2把日期修改为年-月-日的形式

ob = input("请输入学习的科目：")
beizhu = input("请输入备注信息(从哪开始学起？)，并按Enter键开始计时：")


# 开始

time1 = time.time()
date = time.strftime("%Y/%m/%d")
past_time = eval(input("请问你已经学习多少分钟呀："))
start = (datetime.datetime.now() + datetime.timedelta(minutes = (-1)*past_time)).strftime('%H:%M:%S')  ##24小时格式
print('计时开始,按F2结束。没事请不要按F2键啊。否则会停止')

# 结束

keyboard.wait('f2')
time2 = time.time()
end = time.strftime("%H:%M:%S")  ##24小时格式
sum = (time2-time1+past_time*60)/60/60
print("童鞋本次{}学习中，你一共学习了{}小时".format(ob,sum))
sum1 = sum/24

sheet_name = date[0:4] + "年" + date[5:7] + "月"
if not os.path.exists("计时表.xlsx"):
    wb_create = Workbook()
    wb_create.title = sheet_name
    wb_create.save("计时表.xlsx")


wb_load = load_workbook('计时表.xlsx')  # 工作簿
if sheet_name in wb_load:
    ws = wb_load[sheet_name]
else:
    ws = wb_load.create_sheet(sheet_name, 0)

  # 工作簿上的工作表
# 如果不存在计时表.xlsx则创建



if ws.cell(row=1, column=1).value == None:
    ws.cell(row=1, column=1).value = "科目"
    ws.cell(row=1, column=2).value = "备注"
    ws.cell(row=1, column=3).value = "date"
    ws.cell(row=1, column=4).value = "start"
    ws.cell(row=1, column=5).value = "end"
    ws.cell(row=1, column=6).value = "last"
    ws.cell(row=1, column=7).value = "备注"
for i in range(2,1000):
    if ws.cell(row=i, column=1).value == None:
        ws.cell(row=i, column=1).value = ob  # 工作表上的单元格
        ws.cell(row=i, column=2).value = beizhu
        ws.cell(row=i, column=3).value = date
        ws.cell(row=i, column=4).value = start
        ws.cell(row=i, column=5).value = end
        ws.cell(row=i, column=6).value = sum1
        ws.cell(row=i, column=7).value = input("请输入备注信息：")
        break
wb_load.save('计时表.xlsx')